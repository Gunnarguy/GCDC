"""Export all sheets from the GCDC meta spreadsheet XLSX to individual CSVs."""

import csv
import os
import re

import openpyxl

XLSX_PATH = "data/raw/gcdc_meta_spreadsheet/gcdc_meta_spreadsheet.xlsx"
OUT_DIR = "data/raw/gcdc_meta_spreadsheet"


def slugify(name: str) -> str:
    """Turn a sheet name into a safe filename slug."""
    s = name.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def main():
    print(f"Loading {XLSX_PATH} …")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    print(f"Found {len(wb.sheetnames)} sheets:\n")

    os.makedirs(OUT_DIR, exist_ok=True)

    for idx, name in enumerate(wb.sheetnames):
        slug = slugify(name)
        csv_path = os.path.join(OUT_DIR, f"{slug}.csv")
        ws = wb[name]
        row_count = 0
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
                row_count += 1
        print(f"  [{idx:2d}] {name!r:40s} -> {csv_path}  ({row_count} rows)")

    wb.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
